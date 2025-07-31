from flask import Blueprint, jsonify, render_template_string, request
from flask_login import login_required, current_user
from models import db, Party
from datetime import datetime, date
import random
from models import Item, Sale, Purchase
from sqlalchemy import desc

parties_api = Blueprint('parties_api', __name__)

# Parties statistics
@parties_api.route('/api/parties/stats/total')
@login_required
def parties_stats_total():
    """Get total parties count"""
    try:
        count = Party.query.filter_by(user_id=current_user.id).count()
        return str(count)
    except Exception as e:
        print(f"Stats total error: {e}")
        return "0"

@parties_api.route('/api/parties/stats/active')
@login_required
def parties_stats_active():
    """Get active parties count"""
    try:
        count = Party.query.filter_by(user_id=current_user.id).count()  # All parties are considered active
        return str(count)
    except Exception as e:
        print(f"Stats active error: {e}")
        return "0"

@parties_api.route('/api/parties/stats/new')
@login_required
def parties_stats_new():
    """Get new parties count this month"""
    try:
        from datetime import datetime
        start_of_month = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        count = Party.query.filter(
            Party.user_id == current_user.id,
            Party.created_date >= start_of_month
        ).count()
        return str(count)
    except Exception as e:
        print(f"Stats new error: {e}")
        return "0"

# Live search
@parties_api.route('/api/parties/search')
@login_required
def parties_search():
    """Live search parties"""
    try:
        query = request.args.get('q', '').strip()
        if len(query) < 2:
            return jsonify([])
        
        # Use LIKE instead of ILIKE for SQLite compatibility
        search_pattern = f'%{query}%'
        parties = Party.query.filter(
            Party.user_id == current_user.id,
            (Party.party_nm.like(search_pattern) |
             Party.phone.like(search_pattern) |
             Party.mobile.like(search_pattern) |
             Party.place.like(search_pattern) |
             Party.email.like(search_pattern))
        ).limit(10).all()
        
        results = []
        for party in parties:
            results.append({
                'party_cd': party.party_cd,
                'party_nm': party.party_nm,
                'phone': party.phone or party.mobile or 'N/A',
                'place': party.place or 'N/A',
                'email': party.email,
                'balance': party.ytd_dr - party.ytd_cr
            })
        
        return jsonify(results)
    except Exception as e:
        print(f"Search error: {e}")
        return jsonify([])

# Parties table
@parties_api.route('/api/parties/table')
@login_required
def parties_table():
    """Get parties table HTML"""
    try:
        page = request.args.get('page', 1, type=int)
        search = request.args.get('search', '')
        
        query = Party.query.filter_by(user_id=current_user.id)
        
        if search:
            search_pattern = f'%{search}%'
            query = query.filter(
                (Party.party_nm.like(search_pattern) |
                 Party.phone.like(search_pattern) |
                 Party.mobile.like(search_pattern) |
                 Party.place.like(search_pattern))
            )
        
        parties = query.order_by(Party.party_nm).paginate(
            page=page, per_page=20, error_out=False
        )
        
        return render_template_string("""
            {% if parties.items %}
                <div class="table-responsive">
                    <table class="table table-hover mb-0">
                        <thead class="table-dark">
                            <tr>
                                <th>Code</th>
                                <th>Name</th>
                                <th>Contact</th>
                                <th>Location</th>
                                <th>Balance</th>
                                <th>Status</th>
                                <th>Actions</th>
                            </tr>
                        </thead>
                        <tbody>
                            {% for party in parties.items %}
                            <tr>
                                <td>
                                    <span class="badge bg-primary">{{ party.party_cd }}</span>
                                </td>
                                <td>
                                    <strong>{{ party.party_nm }}</strong>
                                    {% if party.party_nm_hindi %}
                                        <br><small class="text-muted">{{ party.party_nm_hindi }}</small>
                                    {% endif %}
                                </td>
                                <td>
                                    {% if party.phone %}
                                        <div><i class="fas fa-phone me-1"></i>{{ party.phone }}</div>
                                    {% endif %}
                                    {% if party.mobile %}
                                        <div><i class="fas fa-mobile-alt me-1"></i>{{ party.mobile }}</div>
                                    {% endif %}
                                    {% if party.email %}
                                        <div><i class="fas fa-envelope me-1"></i>{{ party.email }}</div>
                                    {% endif %}
                                </td>
                                <td>
                                    <div>{{ party.place or 'N/A' }}</div>
                                    {% if party.address1 %}
                                        <small class="text-muted">{{ party.address1 }}</small>
                                    {% endif %}
                                </td>
                                <td>
                                    {% set balance = party.ytd_dr - party.ytd_cr %}
                                    <span class="badge bg-{{ 'danger' if balance > 0 else 'success' }}">
                                        ₹{{ "%.2f"|format(balance) }}
                                    </span>
                                </td>
                                <td>
                                    <span class="badge bg-{{ 'success' if party.bal_cd == 'C' else 'warning' }}">
                                        {{ 'Credit' if party.bal_cd == 'C' else 'Debit' }}
                                    </span>
                                </td>
                                <td>
                                    <div class="btn-group" role="group">
                                        <button class="btn btn-sm btn-outline-primary"
                                                hx-get="/api/parties/view/{{ party.party_cd }}"
                                                hx-target="#partyModal .modal-content"
                                                hx-swap="innerHTML"
                                                data-bs-toggle="modal"
                                                data-bs-target="#partyModal"
                                                title="View Details">
                                            <i class="fas fa-eye"></i>
                                        </button>
                                        <button class="btn btn-sm btn-outline-warning"
                                                hx-get="/api/parties/edit-form/{{ party.party_cd }}"
                                                hx-target="#partyModal .modal-content"
                                                hx-swap="innerHTML"
                                                data-bs-toggle="modal"
                                                data-bs-target="#partyModal"
                                                title="Edit Party">
                                            <i class="fas fa-edit"></i>
                                        </button>
                                        <button class="btn btn-sm btn-outline-danger"
                                                hx-get="/api/parties/delete-confirm/{{ party.party_cd }}"
                                                hx-target="#partyModal .modal-content"
                                                hx-swap="innerHTML"
                                                data-bs-toggle="modal"
                                                data-bs-target="#partyModal"
                                                title="Delete Party">
                                            <i class="fas fa-trash"></i>
                                        </button>
                                    </div>
                                </td>
                            </tr>
                            {% endfor %}
                        </tbody>
                    </table>
                </div>
            {% else %}
                <div class="text-center p-5">
                    <i class="fas fa-users fa-3x text-muted mb-3"></i>
                    <h5 class="text-muted">No parties found</h5>
                    <p class="text-muted">Start by adding your first party</p>
                    <button class="btn btn-primary"
                            hx-get="/api/parties/add-form"
                            hx-target="#partyModal .modal-content"
                            hx-swap="innerHTML"
                            data-bs-toggle="modal"
                            data-bs-target="#partyModal">
                        <i class="fas fa-plus me-2"></i>Add First Party
                    </button>
                </div>
            {% endif %}
        """, parties=parties)
    except Exception as e:
        print(f"Table error: {e}")
        return f"<div class='alert alert-danger'>Error loading parties: {str(e)}</div>"

# Party forms
@parties_api.route('/api/parties/add-form')
@login_required
def parties_add_form():
    """Add party form"""
    try:
        # Generate a default party code
        default_party_code = f'P{current_user.id:02d}{random.randint(100, 999)}'
        
        return render_template_string("""
            <div class="modal-content bright-modal">
                <div class="modal-header bright-header">
                    <h5 class="modal-title">
                        <i class="fas fa-user-plus me-2"></i>Add New Party
                    </h5>
                    <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
                </div>
                <form hx-post="/api/parties/save" hx-target="#partyModal .modal-content" hx-swap="innerHTML">
                    <div class="modal-body bright-body">
                        <div class="row">
                            <div class="col-md-6 mb-3">
                                <label class="form-label bright-label">Party Code *</label>
                                <input type="text" class="form-control bright-input" name="party_cd" required 
                                       value="{{ default_party_code }}">
                            </div>
                            <div class="col-md-6 mb-3">
                                <label class="form-label bright-label">Party Name *</label>
                                <input type="text" class="form-control bright-input" name="party_nm" required>
                            </div>
                        </div>
                        <div class="row">
                            <div class="col-md-6 mb-3">
                                <label class="form-label bright-label">Phone</label>
                                <input type="text" class="form-control bright-input" name="phone">
                            </div>
                            <div class="col-md-6 mb-3">
                                <label class="form-label bright-label">Mobile</label>
                                <input type="text" class="form-control bright-input" name="mobile">
                            </div>
                        </div>
                        <div class="row">
                            <div class="col-md-6 mb-3">
                                <label class="form-label bright-label">Place</label>
                                <input type="text" class="form-control bright-input" name="place">
                            </div>
                            <div class="col-md-6 mb-3">
                                <label class="form-label bright-label">Email</label>
                                <input type="email" class="form-control bright-input" name="email">
                            </div>
                        </div>
                        <div class="row">
                            <div class="col-md-6 mb-3">
                                <label class="form-label bright-label">Opening Balance</label>
                                <input type="number" step="0.01" class="form-control bright-input" 
                                       name="opening_balance" value="0">
                            </div>
                            <div class="col-md-6 mb-3">
                                <label class="form-label bright-label">Balance Type</label>
                                <select class="form-control bright-input" name="bal_cd">
                                    <option value="D">Debit</option>
                                    <option value="C">Credit</option>
                                </select>
                            </div>
                        </div>
                        <div class="mb-3">
                            <label class="form-label bright-label">Address</label>
                            <textarea class="form-control bright-input" name="address1" rows="2"></textarea>
                        </div>
                    </div>
                    <div class="modal-footer bright-footer">
                        <button type="button" class="btn btn-secondary" data-bs-dismiss="modal">Cancel</button>
                        <button type="submit" class="btn btn-primary">
                            <i class="fas fa-save me-2"></i>Save Party
                        </button>
                    </div>
                </form>
            </div>
        """, default_party_code=default_party_code)
    except Exception as e:
        print(f"Add form error: {e}")
        return f"<div class='alert alert-danger'>Error loading form: {str(e)}</div>"

# Party view details
@parties_api.route('/api/parties/view/<party_code>')
@login_required
def parties_view(party_code):
    """View party details"""
    try:
        party = Party.query.filter_by(user_id=current_user.id, party_cd=party_code).first()
        if not party:
            return "<div class='alert alert-danger'>Party not found</div>"
        
        # Calculate current balance
        current_balance = party.opening_bal
        if party.bal_cd == 'D':
            current_balance += (party.ytd_dr or 0) - (party.ytd_cr or 0)
        else:
            current_balance += (party.ytd_cr or 0) - (party.ytd_dr or 0)
        
        return render_template_string("""
            <div class="modal-content bright-modal">
                <div class="modal-header bright-header">
                    <h5 class="modal-title">
                        <i class="fas fa-user me-2"></i>Party Details
                    </h5>
                    <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
                </div>
                <div class="modal-body bright-body">
                    <div class="row">
                        <div class="col-md-6">
                            <h6 class="text-primary mb-3">
                                <i class="fas fa-info-circle me-2"></i>Basic Information
                            </h6>
                            <table class="table table-borderless">
                                <tr>
                                    <td class="fw-bold">Party Code:</td>
                                    <td><span class="badge bg-primary">{{ party.party_cd }}</span></td>
                                </tr>
                                <tr>
                                    <td class="fw-bold">Party Name:</td>
                                    <td>{{ party.party_nm }}</td>
                                </tr>
                                <tr>
                                    <td class="fw-bold">Phone:</td>
                                    <td>{{ party.phone or 'N/A' }}</td>
                                </tr>
                                <tr>
                                    <td class="fw-bold">Mobile:</td>
                                    <td>{{ party.mobile or 'N/A' }}</td>
                                </tr>
                                <tr>
                                    <td class="fw-bold">Email:</td>
                                    <td>{{ party.email or 'N/A' }}</td>
                                </tr>
                            </table>
                        </div>
                        <div class="col-md-6">
                            <h6 class="text-primary mb-3">
                                <i class="fas fa-map-marker-alt me-2"></i>Address & Location
                            </h6>
                            <table class="table table-borderless">
                                <tr>
                                    <td class="fw-bold">Place:</td>
                                    <td>{{ party.place or 'N/A' }}</td>
                                </tr>
                                <tr>
                                    <td class="fw-bold">Address:</td>
                                    <td>{{ party.address1 or 'N/A' }}</td>
                                </tr>
                            </table>
                        </div>
                    </div>
                    
                    <hr class="my-4">
                    
                    <div class="row">
                        <div class="col-12">
                            <h6 class="text-primary mb-3">
                                <i class="fas fa-chart-line me-2"></i>Financial Information
                            </h6>
                            <div class="row">
                                <div class="col-md-3">
                                    <div class="card bg-light">
                                        <div class="card-body text-center">
                                            <h6 class="card-title text-muted">Opening Balance</h6>
                                            <h4 class="text-primary">₹{{ "%.2f"|format(party.opening_bal or 0) }}</h4>
                                            <small class="text-muted">{{ party.bal_cd }}</small>
                                        </div>
                                    </div>
                                </div>
                                <div class="col-md-3">
                                    <div class="card bg-light">
                                        <div class="card-body text-center">
                                            <h6 class="card-title text-muted">YTD Debit</h6>
                                            <h4 class="text-danger">₹{{ "%.2f"|format(party.ytd_dr or 0) }}</h4>
                                        </div>
                                    </div>
                                </div>
                                <div class="col-md-3">
                                    <div class="card bg-light">
                                        <div class="card-body text-center">
                                            <h6 class="card-title text-muted">YTD Credit</h6>
                                            <h4 class="text-success">₹{{ "%.2f"|format(party.ytd_cr or 0) }}</h4>
                                        </div>
                                    </div>
                                </div>
                                <div class="col-md-3">
                                    <div class="card bg-primary text-white">
                                        <div class="card-body text-center">
                                            <h6 class="card-title">Current Balance</h6>
                                            <h4>₹{{ "%.2f"|format(current_balance) }}</h4>
                                        </div>
                                    </div>
                                </div>
                            </div>
                        </div>
                    </div>
                    
                    <hr class="my-4">
                    
                    <div class="row">
                        <div class="col-12">
                            <h6 class="text-primary mb-3">
                                <i class="fas fa-calendar me-2"></i>System Information
                            </h6>
                            <table class="table table-borderless">
                                <tr>
                                    <td class="fw-bold">Created Date:</td>
                                    <td>{{ party.created_date.strftime('%d-%m-%Y %H:%M') if party.created_date else 'N/A' }}</td>
                                </tr>
                                <tr>
                                    <td class="fw-bold">Last Modified:</td>
                                    <td>{{ party.modified_date.strftime('%d-%m-%Y %H:%M') if party.modified_date else 'N/A' }}</td>
                                </tr>
                            </table>
                        </div>
                    </div>
                </div>
                <div class="modal-footer bright-footer">
                    <button type="button" class="btn btn-secondary" data-bs-dismiss="modal">
                        <i class="fas fa-times me-2"></i>Close
                    </button>
                    <button type="button" class="btn btn-primary" 
                            hx-get="/api/parties/edit-form/{{ party.party_cd }}" 
                            hx-target="#partyModal .modal-content" 
                            hx-swap="innerHTML">
                        <i class="fas fa-edit me-2"></i>Edit Party
                    </button>
                </div>
            </div>
        """, party=party, current_balance=current_balance)
    except Exception as e:
        print(f"View party error: {e}")
        return f"<div class='alert alert-danger'>Error: {str(e)}</div>"

# Party edit form
@parties_api.route('/api/parties/edit-form/<party_code>')
@login_required
def parties_edit_form(party_code):
    """Edit party form"""
    try:
        party = Party.query.filter_by(user_id=current_user.id, party_cd=party_code).first()
        if not party:
            return "<div class='alert alert-danger'>Party not found</div>"
        
        return render_template_string("""
            <div class="modal-content bright-modal">
                <div class="modal-header bright-header">
                    <h5 class="modal-title">
                        <i class="fas fa-edit me-2"></i>Edit Party
                    </h5>
                    <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
                </div>
                <form hx-put="/api/parties/update/{{ party.party_cd }}" hx-target="#partyModal .modal-content" hx-swap="innerHTML">
                    <div class="modal-body bright-body">
                        <div class="row">
                            <div class="col-md-6 mb-3">
                                <label class="form-label bright-label">Party Code *</label>
                                <input type="text" class="form-control bright-input" name="party_cd" required 
                                       value="{{ party.party_cd }}" readonly>
                            </div>
                            <div class="col-md-6 mb-3">
                                <label class="form-label bright-label">Party Name *</label>
                                <input type="text" class="form-control bright-input" name="party_nm" required 
                                       value="{{ party.party_nm }}">
                            </div>
                        </div>
                        <div class="row">
                            <div class="col-md-6 mb-3">
                                <label class="form-label bright-label">Phone</label>
                                <input type="text" class="form-control bright-input" name="phone" 
                                       value="{{ party.phone or '' }}">
                            </div>
                            <div class="col-md-6 mb-3">
                                <label class="form-label bright-label">Mobile</label>
                                <input type="text" class="form-control bright-input" name="mobile" 
                                       value="{{ party.mobile or '' }}">
                            </div>
                        </div>
                        <div class="row">
                            <div class="col-md-6 mb-3">
                                <label class="form-label bright-label">Place</label>
                                <input type="text" class="form-control bright-input" name="place" 
                                       value="{{ party.place or '' }}">
                            </div>
                            <div class="col-md-6 mb-3">
                                <label class="form-label bright-label">Email</label>
                                <input type="email" class="form-control bright-input" name="email" 
                                       value="{{ party.email or '' }}">
                            </div>
                        </div>
                        <div class="row">
                            <div class="col-md-6 mb-3">
                                <label class="form-label bright-label">Opening Balance</label>
                                <input type="number" step="0.01" class="form-control bright-input" 
                                       name="opening_balance" value="{{ party.opening_bal or 0 }}">
                            </div>
                            <div class="col-md-6 mb-3">
                                <label class="form-label bright-label">Balance Type</label>
                                <select class="form-control bright-input" name="bal_cd">
                                    <option value="D" {{ 'selected' if party.bal_cd == 'D' else '' }}>Debit</option>
                                    <option value="C" {{ 'selected' if party.bal_cd == 'C' else '' }}>Credit</option>
                                </select>
                            </div>
                        </div>
                        <div class="mb-3">
                            <label class="form-label bright-label">Address</label>
                            <textarea class="form-control bright-input" name="address1" rows="2">{{ party.address1 or '' }}</textarea>
                        </div>
                    </div>
                    <div class="modal-footer bright-footer">
                        <button type="button" class="btn btn-secondary" data-bs-dismiss="modal">Cancel</button>
                        <button type="submit" class="btn btn-primary">
                            <i class="fas fa-save me-2"></i>Update Party
                        </button>
                    </div>
                </form>
            </div>
        """, party=party)
    except Exception as e:
        print(f"Edit form error: {e}")
        return f"<div class='alert alert-danger'>Error loading form: {str(e)}</div>"

# Party update
@parties_api.route('/api/parties/update/<party_code>', methods=['PUT'])
@login_required
def parties_update(party_code):
    """Update existing party"""
    try:
        party = Party.query.filter_by(user_id=current_user.id, party_cd=party_code).first()
        if not party:
            return "<div class='alert alert-danger'>Party not found</div>"
        
        data = request.form
        print(f"Update form data: {dict(data)}")
        
        # Update party fields
        party.party_nm = data.get('party_nm')
        party.phone = data.get('phone')
        party.mobile = data.get('mobile')
        party.email = data.get('email')
        party.place = data.get('place')
        party.address1 = data.get('address1')
        party.opening_bal = float(data.get('opening_balance', 0))
        party.bal_cd = data.get('bal_cd', 'D')
        party.modified_date = datetime.now()
        
        db.session.commit()
        
        print(f"Party updated successfully: {party.party_cd}")
        
        return render_template_string("""
            <div class="text-center">
                <i class="fas fa-check-circle text-success fa-3x mb-3"></i>
                <h5 class="text-success">Party Updated Successfully!</h5>
                <p>Party "{{ party.party_nm }}" ({{ party.party_cd }}) has been updated.</p>
                <button type="button" class="btn btn-primary" data-bs-dismiss="modal">Close</button>
            </div>
        """, party=party)
    except Exception as e:
        print(f"Update party error: {e}")
        return render_template_string("""
            <div class="text-center">
                <i class="fas fa-exclamation-triangle text-danger fa-3x mb-3"></i>
                <h5 class="text-danger">Error!</h5>
                <p>{{ error }}</p>
                <button type="button" class="btn btn-secondary" data-bs-dismiss="modal">Close</button>
            </div>
        """, error=str(e))

# Party delete confirmation
@parties_api.route('/api/parties/delete-confirm/<party_code>')
@login_required
def parties_delete_confirm(party_code):
    """Delete party confirmation"""
    try:
        party = Party.query.filter_by(user_id=current_user.id, party_cd=party_code).first()
        if not party:
            return "<div class='alert alert-danger'>Party not found</div>"
        
        return render_template_string("""
            <div class="modal-content bright-modal">
                <div class="modal-header bright-header">
                    <h5 class="modal-title">
                        <i class="fas fa-exclamation-triangle me-2"></i>Confirm Delete
                    </h5>
                    <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
                </div>
                <div class="modal-body bright-body">
                    <div class="text-center">
                        <i class="fas fa-exclamation-triangle text-warning fa-3x mb-3"></i>
                        <h5 class="text-warning">Delete Party?</h5>
                        <p>Are you sure you want to delete party <strong>"{{ party.party_nm }}"</strong> ({{ party.party_cd }})?</p>
                        <p class="text-muted">This action cannot be undone.</p>
                        
                        <div class="alert alert-info">
                            <strong>Party Details:</strong><br>
                            Name: {{ party.party_nm }}<br>
                            Code: {{ party.party_cd }}<br>
                            Phone: {{ party.phone or 'N/A' }}<br>
                            Place: {{ party.place or 'N/A' }}
                        </div>
                    </div>
                </div>
                <div class="modal-footer bright-footer">
                    <button type="button" class="btn btn-secondary" data-bs-dismiss="modal">
                        <i class="fas fa-times me-2"></i>Cancel
                    </button>
                    <button type="button" class="btn btn-danger" 
                            hx-delete="/api/parties/delete/{{ party.party_cd }}" 
                            hx-target="#partyModal .modal-content" 
                            hx-swap="innerHTML">
                        <i class="fas fa-trash me-2"></i>Delete Party
                    </button>
                </div>
            </div>
        """, party=party)
    except Exception as e:
        print(f"Delete confirm error: {e}")
        return f"<div class='alert alert-danger'>Error: {str(e)}</div>"

# Party delete
@parties_api.route('/api/parties/delete/<party_code>', methods=['DELETE'])
@login_required
def parties_delete(party_code):
    """Delete party"""
    try:
        party = Party.query.filter_by(user_id=current_user.id, party_cd=party_code).first()
        if not party:
            return "<div class='alert alert-danger'>Party not found</div>"
        
        party_name = party.party_nm
        party_code = party.party_cd
        
        db.session.delete(party)
        db.session.commit()
        
        print(f"Party deleted successfully: {party_code}")
        
        return render_template_string("""
            <div class="modal-content bright-modal">
                <div class="modal-header bright-header">
                    <h5 class="modal-title">
                        <i class="fas fa-check-circle me-2"></i>Party Deleted
                    </h5>
                    <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
                </div>
                <div class="modal-body bright-body">
                    <div class="text-center">
                        <i class="fas fa-check-circle text-success fa-3x mb-3"></i>
                        <h5 class="text-success">Party Deleted Successfully!</h5>
                        <p>Party "{{ party_name }}" ({{ party_code }}) has been deleted from your system.</p>
                    </div>
                </div>
                <div class="modal-footer bright-footer">
                    <button type="button" class="btn btn-primary" 
                            data-bs-dismiss="modal"
                            onclick="location.reload()">
                        <i class="fas fa-check me-2"></i>OK
                    </button>
                </div>
            </div>
        """, party_name=party_name, party_code=party_code)
    except Exception as e:
        print(f"Delete party error: {e}")
        return render_template_string("""
            <div class="modal-content bright-modal">
                <div class="modal-header bright-header">
                    <h5 class="modal-title">
                        <i class="fas fa-exclamation-triangle me-2"></i>Error
                    </h5>
                    <button type="button" class="btn-close" data-bs-dismiss="modal"></button>
                </div>
                <div class="modal-body bright-body">
                    <div class="text-center">
                        <i class="fas fa-exclamation-triangle text-danger fa-3x mb-3"></i>
                        <h5 class="text-danger">Delete Failed!</h5>
                        <p>{{ error }}</p>
                    </div>
                </div>
                <div class="modal-footer bright-footer">
                    <button type="button" class="btn btn-secondary" data-bs-dismiss="modal">Close</button>
                </div>
            </div>
        """, error=str(e)) 

# Filtered parties data
@parties_api.route('/api/parties/filtered')
@login_required
def parties_filtered():
    """Get filtered parties data"""
    try:
        # Get filter parameters
        place = request.args.get('place', '')
        balance_type = request.args.get('balanceType', '')
        balance_range = request.args.get('balanceRange', '')
        sort_by = request.args.get('sortBy', 'name')
        
        # Build query
        query = Party.query.filter_by(user_id=current_user.id)
        
        # Apply filters
        if place:
            query = query.filter(Party.place.like(f'%{place}%'))
        
        if balance_type:
            query = query.filter(Party.bal_cd == balance_type)
        
        # Apply sorting
        if sort_by == 'name':
            query = query.order_by(Party.party_nm)
        elif sort_by == 'code':
            query = query.order_by(Party.party_cd)
        elif sort_by == 'place':
            query = query.order_by(Party.place)
        elif sort_by == 'balance':
            query = query.order_by(Party.opening_bal.desc())
        elif sort_by == 'created':
            query = query.order_by(Party.created_date.desc())
        
        parties = query.all()
        
        # Apply balance range filter in Python (if needed)
        if balance_range:
            filtered_parties = []
            for party in parties:
                balance = party.opening_bal or 0
                if balance_range == 'positive' and balance > 0:
                    filtered_parties.append(party)
                elif balance_range == 'negative' and balance < 0:
                    filtered_parties.append(party)
                elif balance_range == 'zero' and balance == 0:
                    filtered_parties.append(party)
            parties = filtered_parties
        
        # Convert to JSON
        result = []
        for party in parties:
            result.append({
                'party_cd': party.party_cd,
                'party_nm': party.party_nm,
                'phone': party.phone,
                'mobile': party.mobile,
                'email': party.email,
                'place': party.place,
                'address1': party.address1,
                'opening_bal': party.opening_bal,
                'bal_cd': party.bal_cd,
                'ytd_dr': party.ytd_dr,
                'ytd_cr': party.ytd_cr,
                'created_date': party.created_date.strftime('%Y-%m-%d') if party.created_date else None,
                'modified_date': party.modified_date.strftime('%Y-%m-%d') if party.modified_date else None
            })
        
        return jsonify(result)
    except Exception as e:
        print(f"Filtered parties error: {e}")
        return jsonify([]) 

# Export parties
@parties_api.route('/api/parties/export')
@login_required
def parties_export():
    """Export parties to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        import io
        
        # Get all parties for current user
        parties = Party.query.filter_by(user_id=current_user.id).order_by(Party.party_nm).all()
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Parties"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            'Party Code', 'Party Name', 'Phone', 'Mobile', 'Email', 
            'Place', 'Address', 'Opening Balance', 'Balance Type',
            'YTD Debit', 'YTD Credit', 'Created Date', 'Modified Date'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Data
        for row, party in enumerate(parties, 2):
            ws.cell(row=row, column=1, value=party.party_cd)
            ws.cell(row=row, column=2, value=party.party_nm)
            ws.cell(row=row, column=3, value=party.phone)
            ws.cell(row=row, column=4, value=party.mobile)
            ws.cell(row=row, column=5, value=party.email)
            ws.cell(row=row, column=6, value=party.place)
            ws.cell(row=row, column=7, value=party.address1)
            ws.cell(row=row, column=8, value=party.opening_bal or 0)
            ws.cell(row=row, column=9, value=party.bal_cd)
            ws.cell(row=row, column=10, value=party.ytd_dr or 0)
            ws.cell(row=row, column=11, value=party.ytd_cr or 0)
            ws.cell(row=row, column=12, value=party.created_date.strftime('%Y-%m-%d') if party.created_date else '')
            ws.cell(row=row, column=13, value=party.modified_date.strftime('%Y-%m-%d') if party.modified_date else '')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        from flask import send_file
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'parties_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        print(f"Export error: {e}")
        return jsonify({'error': str(e)}), 500 

# Items API endpoints
@parties_api.route('/api/items/add-form')
@login_required
def items_add_form():
    """Get add item form"""
    try:
        # Generate next item code
        last_item = Item.query.filter_by(user_id=current_user.id).order_by(desc(Item.it_cd)).first()
        if last_item and last_item.it_cd:
            try:
                last_num = int(last_item.it_cd[1:])  # Remove 'I' prefix and convert to int
                next_num = last_num + 1
            except ValueError:
                next_num = 1
        else:
            next_num = 1
        
        next_item_code = f"I{next_num:05d}"
        
        return render_template_string("""
            <div class="modal-header">
                <h5 class="modal-title" id="quickActionModalLabel">
                    <i class="fas fa-box me-2"></i>Add New Item
                </h5>
                <button type="button" class="btn-close" data-bs-dismiss="modal" aria-label="Close"></button>
            </div>
            <form hx-post="/api/items/add" hx-target="#quickActionModal .modal-content" hx-swap="innerHTML">
                <div class="modal-body">
                    <div class="row g-3">
                        <div class="col-md-6">
                            <label for="it_cd" class="form-label">Item Code *</label>
                            <input type="text" class="form-control" id="it_cd" name="it_cd" value="{{ next_code }}" required>
                        </div>
                        <div class="col-md-6">
                            <label for="it_nm" class="form-label">Item Name *</label>
                            <input type="text" class="form-control" id="it_nm" name="it_nm" required>
                        </div>
                        <div class="col-md-6">
                            <label for="category" class="form-label">Category</label>
                            <input type="text" class="form-control" id="category" name="category">
                        </div>
                        <div class="col-md-6">
                            <label for="unit" class="form-label">Unit</label>
                            <input type="text" class="form-control" id="unit" name="unit" value="PCS">
                        </div>
                        <div class="col-md-6">
                            <label for="rate" class="form-label">Rate</label>
                            <input type="number" step="0.01" class="form-control" id="rate" name="rate" value="0.00">
                        </div>
                        <div class="col-md-6">
                            <label for="stock_qty" class="form-label">Stock Quantity</label>
                            <input type="number" step="0.01" class="form-control" id="stock_qty" name="stock_qty" value="0.00">
                        </div>
                    </div>
                </div>
                <div class="modal-footer">
                    <button type="button" class="btn btn-secondary" data-bs-dismiss="modal">Cancel</button>
                    <button type="submit" class="btn btn-success">
                        <i class="fas fa-save me-1"></i>Save Item
                    </button>
                </div>
            </form>
        """, next_code=next_item_code)
    except Exception as e:
        print(f"Items add form error: {e}")
        return "<div class='modal-body'><div class='alert alert-danger'>Error loading form</div></div>"

@parties_api.route('/api/items/add', methods=['POST'])
@login_required
def items_add():
    """Add new item"""
    try:
        data = request.form
        new_item = Item(
            user_id=current_user.id,
            it_cd=data.get('it_cd'),
            it_nm=data.get('it_nm'),
            category=data.get('category', ''),
            unit=data.get('unit', 'PCS'),
            rate=float(data.get('rate', 0)),
            stock_qty=float(data.get('stock_qty', 0))
        )
        db.session.add(new_item)
        db.session.commit()
        
        return render_template_string("""
            <div class="modal-header">
                <h5 class="modal-title text-success">
                    <i class="fas fa-check-circle me-2"></i>Item Added Successfully
                </h5>
                <button type="button" class="btn-close" data-bs-dismiss="modal" aria-label="Close"></button>
            </div>
            <div class="modal-body">
                <div class="alert alert-success">
                    <i class="fas fa-check-circle me-2"></i>
                    Item "{{ item_name }}" has been added successfully!
                </div>
                <p class="text-muted">The item is now available for use in sales and purchases.</p>
            </div>
            <div class="modal-footer">
                <button type="button" class="btn btn-primary" data-bs-dismiss="modal" onclick="location.reload()">Close</button>
            </div>
        """, item_name=data.get('it_nm'))
    except Exception as e:
        print(f"Items add error: {e}")
        return render_template_string("""
            <div class="modal-header">
                <h5 class="modal-title text-danger">
                    <i class="fas fa-exclamation-triangle me-2"></i>Error
                </h5>
                <button type="button" class="btn-close" data-bs-dismiss="modal" aria-label="Close"></button>
            </div>
            <div class="modal-body">
                <div class="alert alert-danger">
                    <i class="fas fa-exclamation-triangle me-2"></i>
                    Error adding item: {{ error }}
                </div>
            </div>
            <div class="modal-footer">
                <button type="button" class="btn btn-secondary" data-bs-dismiss="modal">Close</button>
            </div>
        """, error=str(e))

# Sales API endpoints - REMOVED: This was conflicting with sales_api.py
# The proper sales add form is now handled by sales_api.py

# Sales add route - REMOVED: This was conflicting with sales_api.py
# The proper sales add functionality is now handled by sales_api.py

# Purchases API endpoints
@parties_api.route('/api/purchases/add-form')
@login_required
def purchases_add_form():
    """Get add purchase form"""
    try:
        # Get parties for dropdown
        parties = Party.query.filter_by(user_id=current_user.id).all()
        # Get items for dropdown
        items = Item.query.filter_by(user_id=current_user.id).all()
        
        return render_template_string("""
            <div class="modal-header">
                <h5 class="modal-title" id="quickActionModalLabel">
                    <i class="fas fa-shopping-cart me-2"></i>New Purchase
                </h5>
                <button type="button" class="btn-close" data-bs-dismiss="modal" aria-label="Close"></button>
            </div>
            <form hx-post="/api/purchases/add" hx-target="#quickActionModal .modal-content" hx-swap="innerHTML">
                <div class="modal-body">
                    <div class="row g-3">
                        <div class="col-md-6">
                            <label for="party_cd" class="form-label">Supplier *</label>
                            <select class="form-select" id="party_cd" name="party_cd" required>
                                <option value="">Select Supplier</option>
                                {% for party in parties %}
                                <option value="{{ party.party_cd }}">{{ party.party_nm }} ({{ party.party_cd }})</option>
                                {% endfor %}
                            </select>
                        </div>
                        <div class="col-md-6">
                            <label for="bill_date" class="form-label">Bill Date *</label>
                            <input type="date" class="form-control" id="bill_date" name="bill_date" value="{{ today }}" required>
                        </div>
                        <div class="col-md-6">
                            <label for="it_cd" class="form-label">Item *</label>
                            <select class="form-select" id="it_cd" name="it_cd" required>
                                <option value="">Select Item</option>
                                {% for item in items %}
                                <option value="{{ item.it_cd }}">{{ item.it_nm }} ({{ item.it_cd }})</option>
                                {% endfor %}
                            </select>
                        </div>
                        <div class="col-md-6">
                            <label for="qty" class="form-label">Quantity *</label>
                            <input type="number" step="0.01" class="form-control" id="qty" name="qty" value="1" required>
                        </div>
                        <div class="col-md-6">
                            <label for="rate" class="form-label">Rate *</label>
                            <input type="number" step="0.01" class="form-control" id="rate" name="rate" value="0.00" required>
                        </div>
                        <div class="col-md-6">
                            <label for="discount" class="form-label">Discount</label>
                            <input type="number" step="0.01" class="form-control" id="discount" name="discount" value="0.00">
                        </div>
                    </div>
                </div>
                <div class="modal-footer">
                    <button type="button" class="btn btn-secondary" data-bs-dismiss="modal">Cancel</button>
                    <button type="submit" class="btn btn-warning">
                        <i class="fas fa-save me-1"></i>Record Purchase
                    </button>
                </div>
            </form>
        """, parties=parties, items=items, today=datetime.now().strftime('%Y-%m-%d'))
    except Exception as e:
        print(f"Purchases add form error: {e}")
        return "<div class='modal-body'><div class='alert alert-danger'>Error loading form</div></div>"

@parties_api.route('/api/purchases/add', methods=['POST'])
@login_required
def purchases_add():
    """Add new purchase"""
    try:
        data = request.form
        # Generate bill number
        last_purchase = Purchase.query.filter_by(user_id=current_user.id).order_by(desc(Purchase.bill_no)).first()
        bill_no = (last_purchase.bill_no + 1) if last_purchase else 1001
        
        new_purchase = Purchase(
            user_id=current_user.id,
            bill_no=bill_no,
            party_cd=data.get('party_cd'),
            it_cd=data.get('it_cd'),
            bill_date=datetime.strptime(data.get('bill_date'), '%Y-%m-%d'),
            qty=float(data.get('qty')),
            rate=float(data.get('rate')),
            discount=float(data.get('discount', 0)),
            sal_amt=float(data.get('qty')) * float(data.get('rate')) - float(data.get('discount', 0))
        )
        db.session.add(new_purchase)
        db.session.commit()
        
        return render_template_string("""
            <div class="modal-header">
                <h5 class="modal-title text-success">
                    <i class="fas fa-check-circle me-2"></i>Purchase Recorded Successfully
                </h5>
                <button type="button" class="btn-close" data-bs-dismiss="modal" aria-label="Close"></button>
            </div>
            <div class="modal-body">
                <div class="alert alert-success">
                    <i class="fas fa-check-circle me-2"></i>
                    Purchase Bill #{{ bill_no }} has been recorded successfully!
                </div>
                <p class="text-muted">Amount: ₹{{ "%.2f"|format(amount) }}</p>
            </div>
            <div class="modal-footer">
                <button type="button" class="btn btn-primary" data-bs-dismiss="modal" onclick="location.reload()">Close</button>
            </div>
        """, bill_no=bill_no, amount=new_purchase.sal_amt)
    except Exception as e:
        print(f"Purchases add error: {e}")
        return render_template_string("""
            <div class="modal-header">
                <h5 class="modal-title text-danger">
                    <i class="fas fa-exclamation-triangle me-2"></i>Error
                </h5>
                <button type="button" class="btn-close" data-bs-dismiss="modal" aria-label="Close"></button>
            </div>
            <div class="modal-body">
                <div class="alert alert-danger">
                    <i class="fas fa-exclamation-triangle me-2"></i>
                    Error recording purchase: {{ error }}
                </div>
            </div>
            <div class="modal-footer">
                <button type="button" class="btn btn-secondary" data-bs-dismiss="modal">Close</button>
            </div>
        """, error=str(e)) 

@parties_api.route('/api/parties/list')
@login_required
def parties_list():
    """Get list of parties for dropdowns"""
    try:
        parties = Party.query.filter_by(user_id=current_user.id).all()
        parties_data = []
        for party in parties:
            parties_data.append({
                'party_cd': party.party_cd,
                'party_nm': party.party_nm,
                'phone': party.phone or '',
                'mobile': party.mobile or '',
                'place': party.place or ''
            })
        
        return jsonify({
            'success': True,
            'parties': parties_data
        })
    except Exception as e:
        print(f"Error getting parties list: {e}")
        return jsonify({
            'success': False,
            'error': str(e),
            'parties': []
        }) 